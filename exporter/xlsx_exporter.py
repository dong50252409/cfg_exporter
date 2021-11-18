# -*- coding: utf-8 -*-

import openpyxl
import helper
from table import Table


class XLSXTable(Table):
    def __init__(self, cfg_obj, file):
        super().__init__(cfg_obj, file)

    def load_column(self):
        book = openpyxl.load_workbook(self.full_filename, read_only=True, data_only=True)
        sheet = book[book.sheetnames[helper.args.default_sheet - 1]]
        rows = list(sheet.iter_rows(values_only=True))
        super()._load_column(rows)


def extension():
    return ".xlsx",
