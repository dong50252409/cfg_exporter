import openpyxl
from cfg_exporter.tables.base.table import Table


class XLSXTable(Table):
    def __init__(self, container, file, args):
        super().__init__(container, file, args)

    def load_table(self):
        book = openpyxl.load_workbook(self.full_filename, read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        book.close()
        self._load_table(rows)
